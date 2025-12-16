import io
import os
import math
import requests
import pandas as pd
from dataclasses import dataclass, field
from typing import Any, List, Optional, Tuple, Iterable, Dict
from enum import Enum


NUMERIC_THRESHOLD_HEADER = 0.25
MIN_TABLE_COLUMN_LEN = 2
MIN_REGION_AREA = 1


class BlockType(str, Enum):
    """Enumeration for the different types a Block can be classified as."""
    NARRATIVE = "narrative"
    HEADER_ONLY = "header-only"
    DATA_ONLY = "data-only"
    COMPLETE_TABULAR = "complete-tabular"
    METADATA = "metadata"
    EMPTY = "empty"
    MERGED = "merged"


@dataclass
class SheetGrid:
    """
    Lightweight 0-based grid wrapper over a sheet's values.

    Attributes
    ----------
    name : str
        Name of the Excel sheet.
    sheet_idx : int
        Index of the sheet in the workbook.
    nrows : int
        Number of used rows.
    ncols : int
        Number of used columns.
    _values : List[List[Any]]
        The raw cell values (row-major).
    merged_spans : List[Tuple[int, int, int, int]], optional
        List of merged cell ranges (r0, r1, c0, c1) inclusive 0-based.
    """
    name: str
    sheet_idx: int
    nrows: int
    ncols: int
    _values: List[List[Any]]
    merged_spans: List[Tuple[int, int, int, int]] = field(default_factory=list)

    def get_value(self, r: int, c: int) -> Any:
        """Safe accessor for grid values."""
        if 0 <= r < self.nrows and 0 <= c < self.ncols:
            return self._values[r][c]
        return None

    def is_empty(self, r: int, c: int) -> bool:
        """Checks if a cell is visually empty (None or whitespace string)."""
        v = self.get_value(r, c)
        if v is None:
            return True
        if isinstance(v, str):
            return len(v.strip()) == 0
        return False


@dataclass
class Block:
    """
    Represents a contiguous non-empty region in an Excel sheet.

    Attributes
    ----------
    idx : int
        Unique identifier for the block in the parsing session.
    sheet_name : str
        Name of the sheet containing this block.
    sheet_idx : int
        Index of the sheet.
    top : int
        Top row index (inclusive).
    bottom : int
        Bottom row index (inclusive).
    left : int
        Left column index (inclusive).
    right : int
        Right column index (inclusive).
    area : int
        Number of cells in the block.
    block_type : BlockType
        Classified type of the block (e.g., TABULAR, NARRATIVE).
    headers : List[str]
        List of column headers strings.
    header_extent : int
        Number of rows detected as headers.
    title : str
        Detected title text for the block.
    stats : dict
        Computed statistics (numeric ratios, dimensions, etc.).
    df : pd.DataFrame
        The parsed DataFrame (if successful).
    """
    idx: int = -1
    sheet_name: str = ""
    sheet_idx: int = 0
    top: int = 0
    bottom: int = 0
    left: int = 0
    right: int = 0
    area: int = 0

    block_type: BlockType = BlockType.NARRATIVE
    headers: Optional[List[str]] = None
    header_extent: int = 0
    title: Optional[str] = None
    stats: dict = field(default_factory=dict)
    df: Optional[pd.DataFrame] = None
    
    # Internal usage for parser logic
    delimiter: str = "excel" 
    merged_into: Optional[int] = None


# ---------------------------------------------------------------------
# ExcelParser Class
# ---------------------------------------------------------------------

class ExcelParser:
    """
    Parses Excel files by detecting contiguous blocks of non-empty cells
    and converting them into structured DataFrames.

    It handles:
    - Spatial segmentation (BFS) to find tables.
    - Merged cell propagation.
    - Statistical header detection.
    - Multi-row header merging.
    - Header borrowing for data-only blocks.

    Attributes
    ----------
    file_path : str
        The local path or URL to the Excel file.
    sheets : List[SheetGrid]
        The loaded sheets content.
    blocks : List[Block]
        The segregated and processed blocks.
    _header_registry : Dict[str, List[Block]]
        Internal registry for header borrowing logic.
    """

    def __init__(self, file_path: str):
        """
        Initializes the ExcelParser.

        Parameters
        ----------
        file_path : str
            The local file path or HTTP/HTTPS URL of the .xls/.xlsx file.
        """
        self.file_path = file_path
        self.sheets: List[SheetGrid] = []
        self.blocks: List[Block] = []
        self._header_registry: Dict[str, List[Block]] = {}
        self.warnings: List[str] = []

    def parse(self) -> List[Block]:
        """
        Executes the full parsing workflow.

        Returns
        -------
        List[Block]
            A list of processed Block objects, potentially containing DataFrames.
        """
        self._fetch_workbook()
        self._segregate_blocks()

        for idx, block in enumerate(self.blocks):
            self._process_block(block, idx)
        
        self._merge_compatible_blocks()

        self._decompose_fused_blocks()

        self._attach_metadata()

        return self.blocks

    def _fetch_workbook(self):
        """
        Fetches file content and populates self.sheets using openpyxl or xlrd.
        """
        # 1. Determine source bytes or path
        content = None
        is_path = False
        
        if self.file_path.lower().startswith("http"):
            response = requests.get(self.file_path)
            response.raise_for_status()
            content = io.BytesIO(response.content)
        else:
            if os.path.exists(self.file_path):
                content = self.file_path
                is_path = True
            else:
                raise FileNotFoundError(f"File not found: {self.file_path}")

        # 2. Open Workbook
        wb_xlsx, wb_xls = None, None
        
        # Try openpyxl (.xlsx) first
        try:
            from openpyxl import load_workbook
            # data_only=True gets values, not formulas
            wb_xlsx = load_workbook(filename=content, data_only=True, read_only=False)
        except Exception:
            # Fallback to xlrd (.xls)
            if not is_path and hasattr(content, 'seek'):
                content.seek(0)
            
            try:
                import xlrd
                if is_path:
                    wb_xls = xlrd.open_workbook(content, formatting_info=True)
                else:
                    wb_xls = xlrd.open_workbook(file_contents=content.read(), formatting_info=True)
            except Exception as e:
                raise RuntimeError(
                    f"Failed to open workbook. Ensure 'openpyxl' (xlsx) or 'xlrd' (xls) is installed. Error: {e}"
                )

        # 3. Convert to SheetGrid objects
        self.sheets = []
        if wb_xlsx:
            self._parse_openpyxl(wb_xlsx)
        elif wb_xls:
            self._parse_xlrd(wb_xls)

    def _parse_openpyxl(self, wb):
        """Internal helper to convert openpyxl workbook to SheetGrids."""
        for idx, ws in enumerate(wb.worksheets):
            title = (ws.title or "").lower()
            if "readme" in title:
                continue

            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            values = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=max_row,
                                                                          min_col=1, max_col=max_col)]
            merged_spans = []
            
            # Propagate merged cells
            for rng in ws.merged_cells.ranges:
                r0, r1 = rng.min_row - 1, rng.max_row - 1
                c0, c1 = rng.min_col - 1, rng.max_col - 1
                merged_spans.append((r0, r1, c0, c1))
                
                tl = values[r0][c0] if 0 <= r0 < max_row and 0 <= c0 < max_col else None
                if tl is not None and (isinstance(tl, str) and tl.strip() == ""):
                    tl = None
                
                if tl is not None:
                    for rr in range(r0, r1 + 1):
                        for cc in range(c0, c1 + 1):
                            current = values[rr][cc]
                            if current is None or (isinstance(current, str) and current.strip() == ""):
                                values[rr][cc] = tl

            nrows_used, ncols_used = self._compute_used_bounds(values)
            trimmed = [row[:ncols_used] for row in values[:nrows_used]]
            
            self.sheets.append(SheetGrid(ws.title, idx, nrows_used, ncols_used, trimmed, merged_spans))

    def _parse_xlrd(self, wb):
        """Internal helper to convert xlrd workbook to SheetGrids."""
        for idx in range(wb.nsheets):
            sh = wb.sheet_by_index(idx)
            title = (sh.name or "").lower()
            if "readme" in title:
                continue

            nrows, ncols = sh.nrows, sh.ncols
            values = [[sh.cell_value(r, c) for c in range(ncols)] for r in range(nrows)]
            merged_spans = []

            if hasattr(sh, "merged_cells"):
                for rlo, rhi, clo, chi in sh.merged_cells:
                    r0, r1 = rlo, rhi - 1
                    c0, c1 = clo, chi - 1
                    merged_spans.append((r0, r1, c0, c1))
                    
                    tl = values[r0][c0] if 0 <= r0 < nrows and 0 <= c0 < ncols else None
                    if tl is not None and (isinstance(tl, str) and tl.strip() == ""):
                        tl = None
                    if tl is not None:
                        for rr in range(r0, r1 + 1):
                            for cc in range(c0, c1 + 1):
                                current = values[rr][cc]
                                if current is None or (isinstance(current, str) and current.strip() == ""):
                                    values[rr][cc] = tl

            nrows_used, ncols_used = self._compute_used_bounds(values)
            trimmed = [row[:ncols_used] for row in values[:nrows_used]]
            
            self.sheets.append(SheetGrid(sh.name, idx, nrows_used, ncols_used, trimmed, merged_spans))

    def _segregate_blocks(self):
        """
        Identify contiguous blocks of non-empty cells in all sheets.
        Applies BFS spatial clustering and minimal enclosure consolidation.
        """
        raw_blocks = []

        for grid in self.sheets:
            visited = [[False] * max(1, grid.ncols) for _ in range(max(1, grid.nrows))]
            
            for r in range(grid.nrows):
                for c in range(grid.ncols):
                    if visited[r][c] or grid.is_empty(r, c):
                        visited[r][c] = True
                        continue

                    # BFS to find island
                    queue = [(r, c)]
                    visited[r][c] = True
                    top = bottom = r
                    left = right = c
                    area = 0

                    while queue:
                        rr, cc = queue.pop()
                        if grid.is_empty(rr, cc):
                            continue
                        area += 1
                        top = min(top, rr)
                        bottom = max(bottom, rr)
                        left = min(left, cc)
                        right = max(right, cc)

                        for dr, dc in self._neighbors8():
                            nr, nc = rr + dr, cc + dc
                            if 0 <= nr < grid.nrows and 0 <= nc < grid.ncols:
                                if not visited[nr][nc]:
                                    visited[nr][nc] = True
                                    if not grid.is_empty(nr, nc):
                                        queue.append((nr, nc))

                    if area >= MIN_REGION_AREA:
                        raw_blocks.append(Block(
                            idx=-1,
                            sheet_name=grid.name,
                            sheet_idx=grid.sheet_idx,
                            top=top, bottom=bottom, left=left, right=right,
                            area=area
                        ))

        # Consolidate Enclosures & Sort
        consolidated = self._consolidate_enclosures_minimal(raw_blocks)
        
        # Sort: Sheet, then visual reading order (top-left)
        consolidated.sort(key=lambda b: (b.sheet_idx, b.left, b.top, b.right, -b.area))
        
        # Assign IDs
        for i, blk in enumerate(consolidated):
            blk.idx = i
            
        self.blocks = consolidated

    def _process_block(self, block: Block, idx: int):
        """
        Main processing logic for a single block.
        
        Computes statistics, detects headers, classifies the block,
        and generates the DataFrame if applicable.
        """
        grid = self._get_sheet_by_name(block.sheet_name)
        if not grid:
            return

        # 0. Gate: Minimum distinct columns check
        col_len = self._block_column_len(block, grid)
        if col_len < MIN_TABLE_COLUMN_LEN:
            block.block_type = BlockType.NARRATIVE
            return

        # 1. Compute Profiles (Stats)
        block.stats = self._compute_statistics(block, grid)

        # 2. Detect Header Extent & Title
        hdr_info = self._detect_header_extent(block, grid, block.stats)
        block.header_extent = hdr_info.get("header_extent", 0)
        block.title = self._extract_title_text(block, grid, hdr_info)

        # 3. Extract & Merge Headers
        merged_headers = self._extract_and_merge_headers(block, grid, hdr_info)
        block.headers = merged_headers

        # 4. Classify
        block.block_type = self._classify_block(block, block.stats, merged_headers)

        # 5. Register Header-Bearing Blocks (for borrowing)
        if block.block_type in (BlockType.COMPLETE_TABULAR, BlockType.HEADER_ONLY) and merged_headers:
            self._header_registry.setdefault(block.sheet_name, []).append(block)

        # 6. Parse Data / Handle Borrowing
        if block.block_type == BlockType.DATA_ONLY:
            # Attempt to borrow headers from a previous block on the same sheet
            candidates = self._header_registry.get(block.sheet_name, [])
            donor = self._find_borrow_candidate(block, candidates)
            
            if donor and donor.headers:
                borrowed = self._borrow_headers(block, donor)
                block.headers = borrowed
                block.block_type = BlockType.COMPLETE_TABULAR
                
                # Adjust header info for generation (data starts immediately after title)
                hdr_info_borrow = dict(hdr_info)
                hdr_info_borrow["header_extent"] = 0 
                block.df = self._generate_df(block, grid, borrowed, hdr_info_borrow)
            elif block.stats.get("mean_row_numeric_ratio", 0) > 0.25:
                width = block.right - block.left + 1
                # Generate generic headers (e.g. Column_1, Column_2)
                default_headers = [f"Column_{i+1}" for i in range(width)]
                block.headers = default_headers
                block.block_type = BlockType.COMPLETE_TABULAR
                
                # Generate DF (header extent is 0 as we fabricated them)
                hdr_info_default = {"title_row": None, "header_extent": 0}
                block.df = self._generate_df(block, grid, default_headers, hdr_info_default)
            else:
                block.df = None
        
        elif block.block_type == BlockType.COMPLETE_TABULAR and merged_headers:
            block.df = self._generate_df(block, grid, merged_headers, hdr_info)
        
        else:
            block.df = None

    # -----------------------------------------------------------------
    # Core Logic Methods
    # -----------------------------------------------------------------

    def _compute_statistics(self, block: Block, grid: SheetGrid) -> dict:
        """Computes numeric density and non-empty counts for rows/cols."""
        row_nonempty, row_numeric = [], []
        
        for r in range(block.top, block.bottom + 1):
            vals = [grid.get_value(r, c) for c in range(block.left, block.right + 1)]
            nonempties = [v for v in vals if not self._is_val_empty(v)]
            k = len(nonempties)
            row_nonempty.append(k)
            
            if k == 0:
                row_numeric.append(0.0)
            else:
                ratio = sum(1 for v in nonempties if self._is_numeric_cell(v)) / k
                row_numeric.append(ratio)

        # Basic aggregate stats
        mean_nonempty = sum(row_nonempty) / len(row_nonempty) if row_nonempty else 0
        mean_numeric = sum(row_numeric) / len(row_numeric) if row_numeric else 0

        return {
            "row_nonempty_counts": row_nonempty,
            "row_numeric_ratio": row_numeric,
            "mean_row_nonempty": mean_nonempty,
            "mean_row_numeric_ratio": mean_numeric
        }

    def _detect_header_extent(self, block: Block, grid: SheetGrid, stats: dict) -> dict:
        """
        Determines the title row and how many lines constitute the header.
        
        Title Rule: First row of block has exactly 1 non-empty cell.
        Header Rule: Extend while row numeric ratio < threshold.
        """
        row_nonempty = stats["row_nonempty_counts"]
        row_numeric = stats["row_numeric_ratio"]
        
        title_row_idx = None
        # Check if first row looks like a title (single cell at top-left of block)
        if row_nonempty and row_nonempty[0] == 1:
            if not grid.is_empty(block.top, block.left):
                title_row_idx = block.top

        start_offset = 1 if title_row_idx is not None else 0
        header_extent = 0
        
        for i in range(start_offset, len(row_numeric)):
            if row_numeric[i] < NUMERIC_THRESHOLD_HEADER:
                header_extent += 1
            else:
                break
                
        return {"title_row": title_row_idx, "header_extent": header_extent}

    def _extract_title_text(self, block: Block, grid: SheetGrid, hdr_info: dict) -> Optional[str]:
        tr = hdr_info.get("title_row")
        if tr is None:
            return None
        # Find the single non-empty cell
        for c in range(block.left, block.right + 1):
            val = grid.get_value(tr, c)
            if not self._is_val_empty(val):
                return str(val).strip()
        return None

    def _extract_and_merge_headers(self, block: Block, grid: SheetGrid, hdr_info: dict) -> Optional[List[str]]:
        """
        Extracts multi-row headers, respecting merged cells, and merges them vertically.
        """
        H = hdr_info.get("header_extent", 0)
        if H <= 0:
            return None

        # 1. Build lookup for merged cells
        merged_idx = {}
        for (r0, r1, c0, c1) in grid.merged_spans:
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    merged_idx[(r, c)] = (r0, r1, c0, c1)

        # 2. Extract spans for each header row
        title_row = hdr_info.get("title_row")
        start_row = block.top + (1 if title_row is not None else 0)
        header_rows_spans = []

        for r in range(start_row, start_row + H):
            spans = []
            c = block.left
            while c <= block.right:
                txt = self._norm_text(grid.get_value(r, c))
                if not txt:
                    c += 1
                    continue
                
                # Determine horizontal span of this text
                if (r, c) in merged_idx:
                    # Use merged cell info
                    _, _, c0, c1 = merged_idx[(r, c)]
                    # Clip to block bounds
                    c0, c1 = max(c0, block.left), min(c1, block.right)
                    if c == c0: # Add only once
                        spans.append({"c0": c0, "c1": c1, "text": txt})
                    c = c1 + 1
                else:
                    # Collapse repeats if not strictly merged
                    c0, c1 = c, c
                    while c1 + 1 <= block.right and self._norm_text(grid.get_value(r, c1 + 1)) == txt:
                        c1 += 1
                    spans.append({"c0": c0, "c1": c1, "text": txt})
                    c = c1 + 1
            header_rows_spans.append(spans)

        # 3. Merge vertically down columns
        width = block.right - block.left + 1
        col_tokens = [[] for _ in range(width)]

        for row_spans in header_rows_spans:
            for j in range(block.left, block.right + 1):
                # Check which span covers column j
                token = ""
                for sp in row_spans:
                    if sp["c0"] <= j <= sp["c1"]:
                        token = sp["text"]
                        break
                if token:
                    col_tokens[j - block.left].append(token)

        # 4. Join and Dedupe
        headers = []
        for tokens in col_tokens:
            cleaned = []
            prev = None
            for t in tokens:
                t = t.strip()
                if t and t != prev:
                    cleaned.append(t)
                    prev = t
            headers.append(" ".join(cleaned))

        # 5. Ensure uniqueness
        return self._ensure_unique(headers)

    def _classify_block(self, block: Block, stats: dict, headers: Optional[List[str]]) -> BlockType:
        """Determines BlockType based on header existence and data presence."""
        H = block.header_extent
        # Data starts after title + header
        offset = (1 if block.title else 0) + H
        data_start_row_idx = offset  # relative to stats array 0-index

        # Count data rows (rows with >= 1 non-empty cell)
        data_rows = 0
        counts = stats["row_nonempty_counts"]
        if data_start_row_idx < len(counts):
            for i in range(data_start_row_idx, len(counts)):
                if counts[i] > 0:
                    data_rows += 1

        if headers and data_rows > 0:
            return BlockType.COMPLETE_TABULAR
        if headers and data_rows == 0:
            return BlockType.HEADER_ONLY
        if not headers and data_rows > 0:
            return BlockType.DATA_ONLY
        
        return BlockType.NARRATIVE

    def _generate_df(self, block: Block, grid: SheetGrid, headers: List[str], hdr_info: dict) -> pd.DataFrame:
        """Creates DataFrame from the block's data region."""
        H = hdr_info.get("header_extent", 0)
        tr = hdr_info.get("title_row")
        data_start = block.top + (1 if tr is not None else 0) + H
        
        rows = []
        ncols = len(headers)
        
        for r in range(data_start, block.bottom + 1):
            raw_vals = [grid.get_value(r, c) for c in range(block.left, block.right + 1)]
            
            # Skip completely empty rows
            if all(self._is_val_empty(v) for v in raw_vals):
                continue
                
            # Normalize length
            if len(raw_vals) < ncols:
                raw_vals += [None] * (ncols - len(raw_vals))
            elif len(raw_vals) > ncols:
                raw_vals = raw_vals[:ncols]
            
            rows.append(raw_vals)
            
        return pd.DataFrame(rows, columns=headers)

    # -----------------------------------------------------------------
    # Helpers: Borrowing & Utils
    # -----------------------------------------------------------------

    def _find_borrow_candidate(self, block: Block, candidates: List[Block]) -> Optional[Block]:
        """Finds a preceding header block with exact horizontal alignment."""
        for cand in reversed(candidates):
            if cand.left == block.left and cand.right == block.right:
                return cand
        return None

    def _borrow_headers(self, block: Block, donor: Block) -> List[str]:
        """Copies headers from donor, adjusting length if necessary."""
        width = block.right - block.left + 1
        hdrs = list(donor.headers)
        if len(hdrs) < width:
            hdrs += [f"unnamed_{i}" for i in range(len(hdrs), width)]
        elif len(hdrs) > width:
            hdrs = hdrs[:width]
        return hdrs

    def _get_sheet_by_name(self, name: str) -> Optional[SheetGrid]:
        for s in self.sheets:
            if s.name == name:
                return s
        return None

    def _block_column_len(self, block: Block, grid: SheetGrid) -> int:
        """Counts columns that have at least one non-empty cell."""
        count = 0
        for c in range(block.left, block.right + 1):
            for r in range(block.top, block.bottom + 1):
                if not grid.is_empty(r, c):
                    count += 1
                    break
        return count

    # -----------------------------------------------------------------
    # Static Utils
    # -----------------------------------------------------------------

    @staticmethod
    def _neighbors8() -> Iterable[Tuple[int, int]]:
        return [(-1, -1), (-1, 0), (-1, 1),
                (0, -1),           (0, 1),
                (1, -1),  (1, 0),  (1, 1)]

    @staticmethod
    def _compute_used_bounds(values: List[List[Any]]) -> Tuple[int, int]:
        """Trims trailing empty rows and columns."""
        if not values:
            return 0, 0
        nrows = len(values)
        ncols = max((len(row) for row in values), default=0)

        # Pad rows
        for row in values:
            if len(row) < ncols:
                row += [None] * (ncols - len(row))

        def is_empty(v):
            return v is None or (isinstance(v, str) and not v.strip())

        last_row = -1
        for r in range(nrows - 1, -1, -1):
            if any(not is_empty(values[r][c]) for c in range(ncols)):
                last_row = r
                break
        
        if last_row < 0:
            return 0, 0

        last_col = -1
        for c in range(ncols - 1, -1, -1):
            if any(not is_empty(values[r][c]) for r in range(last_row + 1)):
                last_col = c
                break
        
        return last_row + 1, last_col + 1

    @staticmethod
    def _consolidate_enclosures_minimal(blocks: List[Block]) -> List[Block]:
        """Removes blocks fully enclosed within larger blocks on the same sheet."""
        by_sheet = {}
        for b in blocks:
            by_sheet.setdefault(b.sheet_idx, []).append(b)

        survivors = []
        for _, sheet_blocks in by_sheet.items():
            # Sort by area descending
            sorted_blocks = sorted(sheet_blocks, key=lambda b: b.area, reverse=True)
            dropped = set()
            
            for i, A in enumerate(sorted_blocks):
                if i in dropped: continue
                for j, B in enumerate(sorted_blocks):
                    if i == j or j in dropped: continue
                    # Check if B inside A
                    if (A.left <= B.left and B.right <= A.right and 
                        A.top <= B.top and B.bottom <= A.bottom):
                        dropped.add(j)
            
            survivors.extend([b for k, b in enumerate(sorted_blocks) if k not in dropped])
        return survivors

    @staticmethod
    def _is_numeric_cell(val: Any) -> bool:
        if val is None: return False
        if isinstance(val, (int, float)):
            return not (isinstance(val, float) and math.isnan(val))
        if isinstance(val, str):
            s = val.strip().lstrip("([{").rstrip(")]}")
            if not s: return False
            if s.endswith("%"):
                try: 
                    float(s[:-1])
                    return True
                except ValueError: pass
            try:
                float(s.replace(",", ""))
                return True
            except ValueError:
                return False
        return False

    @staticmethod
    def _is_val_empty(v: Any) -> bool:
        return v is None or (isinstance(v, str) and not v.strip())

    @staticmethod
    def _norm_text(v: Any) -> str:
        if v is None: return ""
        return str(v).strip()

    @staticmethod
    def _ensure_unique(names: List[str]) -> List[str]:
        seen = {}
        out = []
        for n in names:
            base = n if n else "unnamed"
            if base not in seen:
                seen[base] = 1
                out.append(base)
            else:
                seen[base] += 1
                out.append(f"{base}_{seen[base]}")
        return out
    
    def _merge_compatible_blocks(self):
        """
        Post-processing step to stitch together blocks that are likely 
        parts of the same table (split by empty columns).
        
        Criteria:
        1. Same Sheet
        2. Same Vertical Span (Top/Bottom rows)
        3. Same Header Extent (Structure matches)
        """
        # 1. Group blocks by unique vertical footprint
        # Key: (sheet_idx, top, bottom)
        groups = {}
        for block in self.blocks:
            # Skip if block failed to produce data or is not relevant
            if block.df is None or block.df.empty:
                continue
            
            key = (block.sheet_idx, block.top, block.bottom)
            groups.setdefault(key, []).append(block)

        # 2. Iterate through groups and merge horizontally
        for key, group in groups.items():
            if len(group) < 2:
                continue

            # Sort by visual reading order (left to right)
            group.sort(key=lambda b: b.left)

            # Iterative Merge (Greedy)
            # 'current_main' is the block collecting the data
            current_main = group[0]

            for i in range(1, len(group)):
                next_block = group[i]

                # Check Compatibility: Headers must align structurally
                if current_main.header_extent == next_block.header_extent:
                    self._execute_merge(current_main, next_block)
                else:
                    # If mismatch, this block becomes the new potential 'main'
                    current_main = next_block

    def _execute_merge(self, main: Block, victim: Block):
        """
        Merges 'victim' into 'main' and updates metadata.
        """
        # 1. Concatenate DataFrames (Horizontal)
        try:
            # Reset index to ensure alignment (just in case)
            df_main = main.df.reset_index(drop=True)
            df_victim = victim.df.reset_index(drop=True)
            
            merged_df = pd.concat([df_main, df_victim], axis=1)
            main.df = merged_df
        except Exception as e:
            self.warnings.append(f"Failed to merge Block {main.idx} and {victim.idx}: {e}")
            return

        # 2. Update Metadata
        if victim.headers:
            main.headers = (main.headers or []) + victim.headers
        
        # Visually expand the main block boundaries
        main.right = max(main.right, victim.right)
        
        # 3. Mark Victim
        victim.block_type = BlockType.MERGED
        victim.merged_into = main.idx
        victim.df = None # Release memory
        
        # 4. Log
        msg = (f"Merged Block {victim.idx} into {main.idx} "
               f"(Sheet: {main.sheet_name}, Rows: {main.top}-{main.bottom})")
        self.warnings.append(msg)

    def _attach_metadata(self):
        """
        Iterates through blocks on each sheet to identify and attach
        narrative context to dataframes.
        
        Rules:
        1. Text before the first table -> 'master_metadata' (applies to all tables in sheet).
        2. Text between Table A and Table B -> 'table_metadata' for Table B.
        """
        # Group by sheet to process sequentially
        by_sheet = {}
        for b in self.blocks:
            # Respect the merge logic: ignore victims
            if b.block_type == BlockType.MERGED:
                continue
            by_sheet.setdefault(b.sheet_idx, []).append(b)

        for _, sheet_blocks in by_sheet.items():
            # Sort strictly by vertical position (top row)
            # If tops are equal, left comes first (reading order)
            sheet_blocks.sort(key=lambda b: (b.top, b.left))

            master_accumulator: List[str] = []
            local_accumulator: List[str] = []
            first_table_found = False

            for block in sheet_blocks:
                is_table = (block.block_type == BlockType.COMPLETE_TABULAR and 
                            block.df is not None and not block.df.empty)

                if is_table:
                    first_table_found = True
                    
                    # Attach collected metadata
                    block.df.attrs["master_metadata"] = "\n".join(master_accumulator)
                    block.df.attrs["table_metadata"] = "\n".join(local_accumulator)
                    
                    # Reset local accumulator (consumed by this table)
                    local_accumulator = []
                
                else:
                    # It's a narrative/text block
                    text = self._get_block_text(block)
                    if not text:
                        continue
                        
                    if not first_table_found:
                        # Before first table -> Master
                        master_accumulator.append(text)
                    else:
                        # After first table -> Local (for the NEXT table)
                        local_accumulator.append(text)

    def _get_block_text(self, block: Block) -> str:
        """Helper to extract and join all text content from a block."""
        grid = self._get_sheet_by_name(block.sheet_name)
        if not grid:
            return ""

        lines = []
        for r in range(block.top, block.bottom + 1):
            row_text = []
            for c in range(block.left, block.right + 1):
                val = grid.get_value(r, c)
                if not self._is_val_empty(val):
                    row_text.append(str(val).strip())
            
            if row_text:
                lines.append(" ".join(row_text))
        
        return "\n".join(lines)

    def _decompose_fused_blocks(self):
        """
        Detects and splits 'fused' blocks based on Column Length Continuity
        and Header Repetition.
        """
        new_blocks = []
        
        # Operate on a copy to allow safe modification/appending
        for block in list(self.blocks):
            if block.block_type != BlockType.COMPLETE_TABULAR or block.df is None or block.df.empty:
                continue
            if block.block_type == BlockType.MERGED:
                continue

            # 1. Calculate profile: (Valid Length) for each column
            col_lengths = self._get_column_lengths(block.df)
            
            # 2. Group columns into contiguous spans of similar length
            # Spans are tuples: (start_col_idx, end_col_idx, median_length)
            spans = self._group_columns_by_length(col_lengths)
            
            if len(spans) <= 1:
                # No structural discontinuity found -> Keep as is
                continue

            # 3. Analyze Spans for Repetition
            valid_sub_blocks = []
            current_span_group = [spans[0]] # Accumulator for the current block
            
            # Identify the headers of the *very first* span in this block 
            # to serve as the reference for pattern matching.
            ref_start, ref_end, _ = spans[0]
            ref_headers = list(block.df.columns[ref_start : ref_end + 1])
            
            for i in range(1, len(spans)):
                curr_span = spans[i]
                start_curr, end_curr, _ = curr_span
                curr_headers = list(block.df.columns[start_curr : end_curr + 1])
                
                # Compare current headers against the reference (first span of current group)
                # Note: We compare against the START of the current accumulation group.
                group_start, group_end, _ = current_span_group[0]
                group_headers = list(block.df.columns[group_start : group_end + 1])
                
                if self._headers_match(group_headers, curr_headers):
                    # REPETITION DETECTED -> Split!
                    # 1. Flush the current accumulator as a block
                    valid_sub_blocks.append(self._create_sub_block(block, current_span_group))
                    # 2. Start new accumulator
                    current_span_group = [curr_span]
                else:
                    # NO REPETITION -> Appendage (e.g. Notes column with different length)
                    current_span_group.append(curr_span)
            
            # Flush the final group
            if current_span_group:
                valid_sub_blocks.append(self._create_sub_block(block, current_span_group))
            
            # 4. Apply Updates
            if len(valid_sub_blocks) > 0:
                # The first sub-block replaces the original block
                first_b = valid_sub_blocks[0]
                block.df = first_b.df
                block.headers = first_b.headers
                block.right = first_b.right # Update geometry
                # Note: 'top'/'bottom' of original block remain to preserve vertical sorting order

                # Append the rest as new blocks
                for b in valid_sub_blocks[1:]:
                    # Ensure unique ID
                    b.idx = len(self.blocks) + len(new_blocks) + 5000 
                    new_blocks.append(b)

        self.blocks.extend(new_blocks)

    def _get_column_lengths(self, df: pd.DataFrame) -> List[int]:
        """Returns the row index of the last non-empty value for each column."""
        lengths = []
        for col in df.columns:
            # Find last valid index. 
            non_na = df[col].notna() & (df[col].astype(str).str.strip() != "")
            if not non_na.any():
                lengths.append(0)
            else:
                # Get integer location of last True
                # iloc/idxmax trick for last true value
                last_idx = non_na[::-1].idxmax() 
                if isinstance(last_idx, int):
                    lengths.append(last_idx + 1)
                else:
                    lengths.append(len(df)) 
        return lengths

    def _group_columns_by_length(self, lengths: List[int]) -> List[Tuple[int, int, int]]:
        """
        Groups contiguous columns into spans based on length continuity.
        Skips/Breaks on length 0 (Gaps).
        Returns: [(start, end, max_len), ...]
        """
        spans = []
        if not lengths:
            return spans

        start = 0
        current_len = lengths[0]
        
        for i in range(1, len(lengths)):
            l = lengths[i]
            
            is_gap = (l == 0)
            was_gap = (current_len == 0)
            
            # Tolerance: allow small variance (e.g. +/- 2 rows)
            diff = abs(l - current_len)
            is_same_len = diff <= 2
            
            if is_gap:
                if not was_gap:
                    # Close previous span
                    spans.append((start, i - 1, max(lengths[start:i])))
                start = i
                current_len = 0
            elif was_gap:
                # Start new span from gap
                start = i
                current_len = l
            elif not is_same_len:
                # Length Jump -> Break
                spans.append((start, i - 1, max(lengths[start:i])))
                start = i
                current_len = l
            # Else: Continue span
            
        # Close final span
        if lengths[start] > 0: 
            spans.append((start, len(lengths) - 1, max(lengths[start:])))
            
        return spans

    def _headers_match(self, h1: List[str], h2: List[str]) -> bool:
        """Checks if two header lists imply repetition."""
        if not h1 or not h2:
            return False
            
        def clean(h):
            s = str(h).lower().strip()
            # Remove '.1', '.2' suffixes
            if "." in s and s.rsplit(".", 1)[1].isdigit():
                s = s.rsplit(".", 1)[0]
            return s

        c1 = [clean(x) for x in h1 if "unnamed" not in str(x).lower()]
        c2 = [clean(x) for x in h2 if "unnamed" not in str(x).lower()]
        
        if not c1 or not c2:
            return False

        # Compare first valid header (e.g. 'Depth')
        return c1[0] == c2[0]

    def _create_sub_block(self, parent: Block, span_group: List[Tuple[int, int, int]]) -> Block:
        """Creates a new block from a list of column spans, cropping rows vertically."""
        # 1. Determine Horizontal Boundaries
        first_span = span_group[0]
        last_span = span_group[-1]
        
        start_col_idx = first_span[0]
        end_col_idx = last_span[1]
        
        # 2. Slice DataFrame Columns
        # Note: This slice inherently includes any "Gap" columns that existed between spans 
        # within this group (e.g. Data | Gap | Notes). This is usually desired if they are grouped.
        sub_df = parent.df.iloc[:, start_col_idx : end_col_idx + 1]
        
        # 3. Determine Vertical Boundary (Crop)
        # Use the max length of the spans to crop empty rows at bottom
        max_h = max(s[2] for s in span_group)
        if max_h < len(sub_df):
            sub_df = sub_df.iloc[:max_h, :]
            
        # 4. Clean boundaries (remove the gaps if we included them by range slicing)
        sub_df = self._clean_boundary_columns(sub_df)

        return Block(
            sheet_name=parent.sheet_name,
            sheet_idx=parent.sheet_idx,
            top=parent.top,
            bottom=parent.top + max_h, # Adjusted bottom
            left=parent.left + start_col_idx, # Relative offset
            right=parent.left + end_col_idx,
            block_type=BlockType.COMPLETE_TABULAR,
            df=sub_df,
            headers=list(sub_df.columns),
            stats=parent.stats,
            title=parent.title
        )

    def _clean_boundary_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Removes unnamed/empty columns from the start/end of a split chunk."""
        if df.empty: return df
        df = df.dropna(axis=1, how='all')
        to_drop = []
        for col in df.columns:
            c_str = str(col).lower()
            if "unnamed" in c_str or c_str == "":
                 if df[col].count() == 0:
                     to_drop.append(col)
        return df.drop(columns=to_drop)


if __name__ == "__main__":
    # Example usage
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/NonStandardParser/Correspondence/notebook/frank1999.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-ocean99-xls/Clemens/Clemens1996/clemens1996.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-ocean99-xls/Ishiwatari/ishiwatari1999.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-ocean99-xls/Overpeck1996/overpeck1996.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-ocean99-xls/Bond/bond1992.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-ocean99-xls/Charles/charles1996.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/NonStandardParser/Correspondence/notebook/marinduque2004.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-coral99-xls/Linsley/Rarotonga/Rarotonga2006/rarotonga2006.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-coral99-xls/DeLong/Delong2007/amedee2007.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-ice99-xls/Schaefer/Pakitsoq2009/pakitsoq2009.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/ExcelParser/Data/orig-ice99-xls/Thompson/Dasuopu/dasuopu2000.xls")
    # parser = ExcelParser("/Users/dhirenoswal/Desktop/TU corpus/SST_Indian_Ocean_Bard1997.xlsx")
    blocks = parser.parse()

    for block in blocks:
        print(f"Block ID: {block.idx}, Type: {block.block_type}, Sheet: {block.sheet_name}")
        if block.df is not None and not block.df.empty:
            # if block.sheet == "":
                print(block.df.head())
                print(block.df.shape)
                print(block.df.attrs)

    # block = blocks[3]
    # print(f"Block ID: {block.idx}, Type: {block.block_type}, Sheet: {block.sheet_name}, Header Extent:{block.header_extent}, Title: {block.title}")
    # block = blocks[17]
    # print(f"Block ID: {block.idx}, Type: {block.block_type}, Sheet: {block.sheet_name}, Header Extent:{block.header_extent}, Title: {block.title}")

    # block = blocks[18]
    # print(f"Block ID: {block.idx}, Type: {block.block_type}, Sheet: {block.sheet_name}, Header Extent:{block.header_extent}, Title: {block.title}")